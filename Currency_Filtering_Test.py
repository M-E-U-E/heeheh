<<<<<<< HEAD
=======
import os
import logging
>>>>>>> 848dcf7 (Add files via upload)
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
<<<<<<< HEAD
from webdriver_manager.chrome import ChromeDriverManager
=======
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')
>>>>>>> 848dcf7 (Add files via upload)

# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
<<<<<<< HEAD
    driver.implicitly_wait(10)  # Wait for elements before raising exceptions
    return driver

# Test: Check if Property Tiles Currency Changes Based on Dropdown Selection
def check_currency_change(driver, url, output_xlsx, output_summary_xlsx):
    driver.get(url)
    time.sleep(2)

    # Attempt to locate the currency selector button and click it
    try:
        # Using WebDriverWait to ensure the element is visible and clickable
        print("Waiting for the currency selector button...")
        currency_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "currency-selector"))  # Update with correct ID if necessary
        )
        print("Currency selector button found, clicking...")
        currency_button.click()
    except Exception as e:
        print(f"Error: Could not locate or click currency selector button. {e}")
        return

    # Wait for the dropdown to become visible
    try:
        print("Waiting for the currency dropdown to appear...")
        currency_dropdown = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, ".currency-selector__list"))  # The class for the dropdown
        )
        print("Currency dropdown found!")
    except Exception as e:
        print(f"Error: Currency dropdown not visible. {e}")
        return

    # Find all currency options in the dropdown
    print("Finding all currency options...")
    currency_options = driver.find_elements(By.CSS_SELECTOR, ".currency-selector__list-item")  # List of options
    print(f"Found {len(currency_options)} currency options.")
    
    for option in currency_options:
        if option.text.strip() == "EUR":  # Example: Choose EUR; can replace with any currency you want
            print("Clicking on EUR...")
            option.click()
            break
    else:
        print("Currency option not found.")
        return

    # Wait for the page to update and verify if the currency in property tiles changed
    time.sleep(3)  # Ensure the page has time to update
    print("Verifying if the currency in property tiles has changed...")
    property_tiles = driver.find_elements(By.CSS_SELECTOR, ".property-tile .price")  # Look for the updated price
    currency_changed = False

    for tile in property_tiles:
        if '€' in tile.text:  # Check for EUR (adjust for other currencies)
            currency_changed = True
            break

    if currency_changed:
        result = "Pass"
        comment = "Currency change verified successfully."
    else:
        result = "Fail"
        comment = "Currency change not reflected in property tiles."

    # Save the result to an Excel file
    test_results = [{
        "URL": url,
        "Testcase": "Currency Filtering Test",
        "Result": result,
        "Comments": comment
    }]
    df_results = pd.DataFrame(test_results)
    df_results.to_excel(output_xlsx, index=False)

    # Create a summary DataFrame for Pass/Fail count
    summary_data = [{
        "Test": "Currency Change Summary",
        "Result": result,
        "Total Currency Options": len(currency_options)
    }]
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(output_summary_xlsx, index=False)

    print(f"Currency change test result saved to {output_xlsx}")
    print(f"Currency change summary saved to {output_summary_xlsx}")

def main():
    url = "https://www.alojamiento.io/"  # Replace with the actual URL
    output_xlsx = "test_results/currency_change_results.xlsx"
    output_summary_xlsx = "test_results/currency_change_summary.xlsx"
    driver = init_driver()

    check_currency_change(driver, url, output_xlsx, output_summary_xlsx)

    driver.quit()
=======
    driver.implicitly_wait(10)
    return driver

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df):
    """
    Save a DataFrame to an Excel file, auto-adjust column widths, and enhance formatting.

    Args:
        filepath (str): Path to save the Excel file.
        df (pd.DataFrame): DataFrame to save.
    """
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5

    # Apply header formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    wb.save(filepath)

# Test currency filter functionality
def test_currency_filter(driver, url):
    logging.info(f"Starting Currency Filter Test for URL: {url}")
    testcase = "Currency Filter Test"
    results = []  # List to store individual test results for each currency

    try:
        driver.get(url)
        logging.info("Page loaded successfully.")

        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # Scroll down to load all content
        for _ in range(3):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)

        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "js-currency-sort-footer"))
        )
        dropdown.click()
        logging.info("Currency dropdown opened.")

        options = dropdown.find_elements(By.CSS_SELECTOR, ".select-ul > li")
        logging.info(f"Found {len(options)} currency options.")

        if not options:
            logging.warning("No currency options found in the dropdown.")
            return [{"Currency Name": "All", "Currency Symbol": "N/A", "Status": "Fail", "Reason": "No currency options found"}]

        for option in options:
            data_country = option.get_attribute("data-currency-country")
            currency_element = option.find_element(By.CSS_SELECTOR, ".option > p")
            currency_symbol = currency_element.text.split(" ")[0].strip()

            try:
                dropdown.click()
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(option)
                ).click()

                tiles = driver.find_elements(By.CLASS_NAME, "js-price-value")
                if not tiles or not all(currency_symbol in tile.text for tile in tiles):
                    results.append({"Currency Name": data_country, "Currency Symbol": currency_symbol, "Status": "Fail", "Reason": "Currency not reflected in tiles"})
                else:
                    results.append({"Currency Name": data_country, "Currency Symbol": currency_symbol, "Status": "Pass", "Reason": "Validation successful"})
            except Exception as e:
                results.append({"Currency Name": data_country, "Currency Symbol": currency_symbol, "Status": "Fail", "Reason": str(e)})
                logging.error(f"Error for currency {currency_symbol}: {str(e)}")

        return results

    except Exception as e:
        logging.error(f"Error during {testcase}: {str(e)}")
        return [{"Currency Name": "All", "Currency Symbol": "N/A", "Status": "Fail", "Reason": f"Exception: {str(e)}"}]

# Main function
def main():
    url = "https://www.alojamiento.io/"  # Replace with the actual URL
    output_dir = "test_results"
    ensure_directory(output_dir)

    output_results_xlsx = os.path.join(output_dir, "currency_test_results.xlsx")
    output_summary_xlsx = os.path.join(output_dir, "currency_test_summary.xlsx")

    driver = init_driver()
    try:
        results = test_currency_filter(driver, url)

        df_results = pd.DataFrame(results)
        save_with_auto_width(output_results_xlsx, df_results)

        fail_results = [res for res in results if res["Status"] == "Fail"]
        pass_count = len([res for res in results if res["Status"] == "Pass"])
        fail_count = len(fail_results)

        overall_status = "Pass" if fail_count == 0 else "Fail"
        comments = "All currencies passed successfully." if fail_count == 0 else f"{fail_count} currencies failed."

        summary_data = [{
            "page_url": url,
            "testcase": "Currency Filter Test",
            "status": overall_status,
            "comments": comments
        }]
        df_summary = pd.DataFrame(summary_data)

        save_with_auto_width(output_summary_xlsx, df_summary)
        logging.info(f"Test results saved to {output_results_xlsx}")
        logging.info(f"Test summary saved to {output_summary_xlsx}")

    except Exception as e:
        logging.error(f"An error occurred during execution: {e}")
    finally:
        driver.quit()
>>>>>>> 848dcf7 (Add files via upload)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
<<<<<<< HEAD
        print("Execution interrupted by user.")
=======
        logging.info("Execution interrupted by user.")
>>>>>>> 848dcf7 (Add files via upload)
