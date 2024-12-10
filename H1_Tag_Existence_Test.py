<<<<<<< HEAD
=======
import os
import logging
>>>>>>> 848dcf7 (Add files via upload)
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
<<<<<<< HEAD
from webdriver_manager.chrome import ChromeDriverManager
=======
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')
>>>>>>> 848dcf7 (Add files via upload)

# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)  # Wait for elements before raising exceptions
    return driver

<<<<<<< HEAD
# Test: Check All H1 Tags and Where They Are Found
def check_all_h1_tags(driver, url):
    driver.get(url)
    time.sleep(2)  # Allow page to load
    h1_tags = driver.find_elements(By.TAG_NAME, "h1")
    
    # If H1 tags are found, return the tag texts and the URL
    if h1_tags:
        h1_texts = [h1.text for h1 in h1_tags]  # Get text of each H1 tag
        return "Pass", "H1 tags found.", h1_texts, url
    
    # If no H1 tags are found
    return "Fail", "No H1 tags found.", [], url

def main():
    url = "https://www.alojamiento.io/"
    
    # Output Excel file paths
    output_xlsx_result = "test_results/h1_tag_results.xlsx"
    output_xlsx_url = "test_results/h1_tag_url_results.xlsx"
    
    driver = init_driver()

    # Run H1 tag test and get all H1 tag texts
    result, comment, h1_texts, found_url = check_all_h1_tags(driver, url)

    # Save the result to the first Excel file (test result)
    test_results = [{
        "page_url": url,
        "testcase": "All H1 Tags Test",
        "result": result,
        "comments": comment,
    }]
    
    df_results = pd.DataFrame(test_results)
    df_results.to_excel(output_xlsx_result, index=False)
    print(f"Test results for All H1 Tags Test saved to {output_xlsx_result}")

    # Save all found H1 tags and URLs to a second Excel file
    if h1_texts:
        h1_tags_data = [{"URL": found_url, "H1 Tag": h1_text} for h1_text in h1_texts]
    else:
        h1_tags_data = [{"URL": found_url, "H1 Tag": "No H1 tags found"}]

    df_h1_tags = pd.DataFrame(h1_tags_data)
    df_h1_tags.to_excel(output_xlsx_url, index=False)
    print(f"All H1 tags saved to {output_xlsx_url}")

    driver.quit()
=======
# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths
def save_with_auto_width(filepath, df):
    """
    Save a DataFrame to an Excel file, auto-adjust column widths, and enhance formatting.

    Args:
        filepath (str): Path to save the Excel file.
        df (pd.DataFrame): DataFrame to save.
    """
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:  # Avoid issues with None values
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5  # Add padding for visibility

    # Apply header formatting
    for cell in ws[1]:  # First row is the header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Save the updated workbook
    wb.save(filepath)

# Test: Check All H1 Tags and Where They Are Found
def check_all_h1_tags(driver, url):
    logging.info(f"Checking H1 tags for URL: {url}")
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2)  # Allow the page to load fully
        h1_tags = driver.find_elements(By.TAG_NAME, "h1")
        
        if h1_tags:
            h1_texts = [h1.text.strip() for h1 in h1_tags if h1.text.strip()]
            logging.info(f"Found {len(h1_texts)} H1 tags on the page.")
            return "Pass", "H1 tags found.", h1_texts
        else:
            logging.warning("No H1 tags found on the page.")
            return "Fail", "No H1 tags found.", []
    except TimeoutException:
        logging.error("Page load timeout.")
        return "Fail", "Page load timeout.", []
    except Exception as e:
        logging.error(f"Error checking H1 tags: {e}")
        return "Fail", f"Error: {e}", []

# Main function
def main():
    url = "https://www.alojamiento.io/property/mall-of-i-stanbul-3/BC-6975002/"  # Replace with the actual URL
    output_dir = "test_results"
    output_xlsx_result = os.path.join(output_dir, "h1_tag_results.xlsx")  # Keep this file unchanged
    output_summary_xlsx = os.path.join(output_dir, "h1_tag_summary.xlsx")  # Create this summary file

    # Ensure the output directory exists
    ensure_directory(output_dir)

    driver = init_driver()
    try:
        # Run the H1 tag test
        result, comment, h1_texts = check_all_h1_tags(driver, url)

        # Save the detailed H1 tag results (unchanged)
        test_results = [{
            "Page URL": url,
            "Test Case": "All H1 Tags Test",
            "Result": result,
            "Comments": comment,
            "Total H1 Tags Found": len(h1_texts)
        }]
        df_results = pd.DataFrame(test_results)
        save_with_auto_width(output_xlsx_result, df_results)
        logging.info(f"Test results saved to {output_xlsx_result}")

        # Generate the summary in the required format
        overall_status = "Pass" if result == "Pass" else "Fail"
        comments = "All H1 tags present." if result == "Pass" else comment

        # Create the summary data
        summary_data = [{
            "page_url": url,
            "testcase": "Test of H1 Tags",
            "status": overall_status,
            "comments": comments
        }]
        df_summary = pd.DataFrame(summary_data)

        # Save the summary to a separate summary file
        save_with_auto_width(output_summary_xlsx, df_summary)
        logging.info(f"Summary saved to {output_summary_xlsx}")

    except Exception as e:
        logging.error(f"Error in main execution: {e}")
    finally:
        driver.quit()
>>>>>>> 848dcf7 (Add files via upload)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
<<<<<<< HEAD
        print("Execution interrupted by user.")
=======
        logging.info("Execution interrupted by user.")
>>>>>>> 848dcf7 (Add files via upload)
