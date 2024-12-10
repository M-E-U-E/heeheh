import os
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df, sheet_name=None):
    df.to_excel(filepath, index=False, engine='openpyxl', sheet_name=sheet_name)
    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    wb.save(filepath)

# Run individual scripts and collect their outputs
def run_tests(test_scripts, result_dir):
    ensure_directory(result_dir)
    for script in test_scripts:
        script_name = os.path.basename(script)
        print(f"Running test: {script_name}")
        try:
            subprocess.run(["python", script], check=True)
        except subprocess.CalledProcessError as e:
            print(f"Error running {script_name}: {e}")

# Consolidate all test results into a single report
def consolidate_results(result_dir, report_file):
    summary_data = []

    for file_name in os.listdir(result_dir):
        if file_name.endswith("_results.xlsx"):
            file_path = os.path.join(result_dir, file_name)
            test_case = file_name.replace("_results.xlsx", "").replace("_", " ").title()

            # Load individual test results
            df = pd.read_excel(file_path)
            if 'testcase' not in df.columns:
                df['testcase'] = test_case
            summary_data.append(df[['page_url', 'testcase', 'result', 'comments']])

            # Add individual test results as a separate sheet
            save_with_auto_width(report_file, df, sheet_name=test_case)

    # Create and save the summary sheet
    if summary_data:
        summary_df = pd.concat(summary_data, ignore_index=True)
        save_with_auto_width(report_file, summary_df, sheet_name="Summary")
        print(f"Consolidated report saved to {report_file}")

# Main function
def main():
    test_scripts = [
        "Currency_Filtering_Test.py",
        "H1_Tag_Existence_Test.py",
        "HTML_Tag_Sequence_Test.py",
        "Image_Alt_Attribute_Test.py",
        "Scrape_Data_from_Script_Tag.py",
        "URL_Status_Code_Test.py",
    ]
    result_dir = "test_results"
    report_file = os.path.join(result_dir, "summary_report.xlsx")

    # Run all tests
    run_tests(test_scripts, result_dir)

    # Consolidate results
    consolidate_results(result_dir, report_file)

if __name__ == "__main__":
    main()
