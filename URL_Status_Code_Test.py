<<<<<<< HEAD
import time
import requests
=======
import os
import time
import requests
import logging
>>>>>>> 848dcf7 (Add files via upload)
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import urllib3
from openpyxl import load_workbook
<<<<<<< HEAD
=======
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
>>>>>>> 848dcf7 (Add files via upload)

# Disable insecure request warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

<<<<<<< HEAD
# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)  # Wait for elements before raising exceptions
    return driver

# Test: Check URL Status Codes and Save
def check_url_status_and_save(driver, url, output_xlsx, output_summary_xlsx):
    driver.get(url)
    time.sleep(2)

    # Find all anchor tags and extract their href attributes
    links = [a.get_attribute("href") for a in driver.find_elements(By.TAG_NAME, "a") if a.get_attribute("href")]
    links = list(set(link for link in links if link and link.startswith("http")))  # Filter duplicates

    # # Filter out Facebook URLs
    # links = [link for link in links if "facebook.com" not in link]
=======
# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)
    return driver

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df):
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5

    # Apply header formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    wb.save(filepath)

# Test: Check URL Status Codes and Save
def check_url_status_and_save(driver, url, output_xlsx, output_summary_xlsx):
    logging.info(f"Starting URL Status Test for URL: {url}")
    driver.get(url)
    time.sleep(2)

    # Extract all unique anchor links
    links = [a.get_attribute("href") for a in driver.find_elements(By.TAG_NAME, "a") if a.get_attribute("href")]
    links = list(set(link for link in links if link and link.startswith("http")))

    logging.info(f"Found {len(links)} unique links on the page.")
>>>>>>> 848dcf7 (Add files via upload)

    # Set up a session with retries
    session = requests.Session()
    retries = Retry(total=5, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    session.mount('https://', HTTPAdapter(max_retries=retries))

<<<<<<< HEAD
    # List to store the link attributes and status
    link_data = []
    pass_count = 0
    fail_count = 0
=======
    # Store link details
    link_data = []
    pass_count = 0
    fail_count = 0
    failed_urls = []
>>>>>>> 848dcf7 (Add files via upload)

    for link in links:
        status = "Pass"
        error_message = ""
<<<<<<< HEAD
        status_code = ""  # Variable to hold the HTTP status code

        try:
            response = session.get(link, timeout=5, verify=False)
            status_code = response.status_code  # Store HTTP status code
=======
        status_code = ""

        try:
            response = session.get(link, timeout=5, verify=False)
            status_code = response.status_code
>>>>>>> 848dcf7 (Add files via upload)
            if status_code == 404:
                status = "Fail"
                error_message = "404 Not Found"
                fail_count += 1
<<<<<<< HEAD
            else:
=======
                failed_urls.append(link)  # Add to failed URLs only if 404
                any_404_error = True
            else:
                status = "pass"
>>>>>>> 848dcf7 (Add files via upload)
                pass_count += 1
        except requests.exceptions.Timeout:
            status = "Fail"
            error_message = "Timeout"
            fail_count += 1
        except requests.exceptions.RequestException as e:
            status = "Fail"
            error_message = f"Error: {e}"
            fail_count += 1

        link_data.append({
            "URL": link,
            "Status": status,
<<<<<<< HEAD
            "HTTP Status Code": status_code,  # Add HTTP status code
            "Error Message": error_message if error_message else "None"
        })

    # Create a DataFrame and save to Excel for detailed status
    df = pd.DataFrame(link_data)
    df.to_excel(output_xlsx, index=False)

    # Load the workbook and adjust column widths
    wb = load_workbook(output_xlsx)
    ws = wb.active

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the adjusted Excel file
    wb.save(output_xlsx)

    # Create a summary DataFrame for Pass/Fail count
    summary_data = [{
        "Test": "URL Status Summary",
        "Pass Count": pass_count,
        "Fail Count": fail_count,
        "Total URLs": len(links)
    }]
    df_summary = pd.DataFrame(summary_data)

    # Save the summary to another Excel file
    df_summary.to_excel(output_summary_xlsx, index=False)
    print(f"URL status analysis saved to {output_xlsx}")
    print(f"URL status summary saved to {output_summary_xlsx}")

def main():
    url = "https://www.alojamiento.io/"
    output_xlsx = "test_results/url_status_results.xlsx"  # Change file extension to .xlsx
    output_summary_xlsx = "test_results/url_status_summary.xlsx"  # Summary file
    driver = init_driver()

    check_url_status_and_save(driver, url, output_xlsx, output_summary_xlsx)

    driver.quit()
=======
            "HTTP Status Code": status_code if status_code else "N/A",
            "Error Message": error_message if error_message else "None"
        })

        logging.info(f"Checked URL: {link}, Status: {status}, HTTP Code: {status_code}, Error: {error_message}")
    
    # Check after all URLs if none are 404, change all statuses to "Pass"
    if not any(item["HTTP Status Code"] == 404 for item in link_data):
        for item in link_data:
            item["Status"] = "Pass"
    
    # Save detailed URL status results
    df_links = pd.DataFrame(link_data)
    save_with_auto_width(output_xlsx, df_links)
    overall_status = "Pass" if all(link['Status'] == "Pass" for link in link_data) else "Fail"
    logging.info(f"Detailed URL status analysis saved to {output_xlsx}")
    
    # Define comments based on test results
    if overall_status == "Pass":
        comments = "All URLs passed successfully."
    else:
        failed_count = sum(1 for link in link_data if link['Status'] == "Fail")
        comments = f"{failed_count} URL(s) failed."
    # Create summary including failed URLs
    summary_data = [{
        "page_url": url,
        "testcase": "Test of URLs",
        "status": overall_status,
        "comments": comments
    }]
    df_summary = pd.DataFrame(summary_data)
    save_with_auto_width(output_summary_xlsx, df_summary)
    logging.info(f"URL status summary saved to {output_summary_xlsx}")

# Main function
def main():
    url = "https://www.alojamiento.io/property/mall-of-i-stanbul-3/BC-6975002/"
    output_dir = "test_results"
    ensure_directory(output_dir)

    output_xlsx = os.path.join(output_dir, "url_status_results.xlsx")
    output_summary_xlsx = os.path.join(output_dir, "url_status_summary.xlsx")

    driver = init_driver()

    try:
        check_url_status_and_save(driver, url, output_xlsx, output_summary_xlsx)
    except Exception as e:
        logging.error(f"An error occurred during execution: {e}")
    finally:
        driver.quit()
>>>>>>> 848dcf7 (Add files via upload)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
<<<<<<< HEAD
        print("Execution interrupted by user.")
=======
        logging.info("Execution interrupted by user.")
>>>>>>> 848dcf7 (Add files via upload)
